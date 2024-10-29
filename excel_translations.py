import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import requests
from io import BytesIO

# DeepL API key
deepl_api_key = '25f727f2-4270-6944-171f-da78e41dcef0:fx'

# Define DeepL languages and their language codes
deepl_languages = {
    "English": {"code": "EN", "column": "B"},
    "French": {"code": "FR", "column": "C"},
    "Dutch": {"code": "NL", "column": "D"},
    "Swedish": {"code": "SV", "column": "E"},
    "Czech": {"code": "CS", "column": "F"},
    "Slovak": {"code": "SK", "column": "G"},
}

# Example glossary entries for each language
glossary_entries_map = {
    "EN": "",
    "FR": "",
    "NL": "",
    "SV": "",
    "CS": "",
    "SK": ""
}
# glossary_entries_map = {
#     "EN": "Welt\tChupapi\nBeispiel\tMunjanju",
#     "FR": "Welt\tChupapi\nBeispiel\tMunjanju",
#     "NL": "Welt\tChupapi\nBeispiel\tMunjanju",
#     "SV": "Welt\tChupapi\nBeispiel\tMunjanju",
#     "CS": "Welt\tChupapi\nBeispiel\tMunjanju",
#     "SK": "Welt\tChupapi\nBeispiel\tMunjanju"
# }

# Function to create a glossary and get its ID for a specific language
def create_glossary(api_key, name, source_lang, target_lang, entries):
    url = "https://api-free.deepl.com/v2/glossaries"
    headers = {'Authorization': f'DeepL-Auth-Key {api_key}'}
    data = {
        'name': name,
        'source_lang': source_lang,
        'target_lang': target_lang,
        'entries': entries,
        'entries_format': 'tsv'
    }
    response = requests.post(url, headers=headers, data=data)
    return response.json().get('glossary_id')

# Create glossaries for each target language
glossary_ids = {}
for language, info in deepl_languages.items():
    entries = glossary_entries_map[info["code"]]
    glossary_id = create_glossary(deepl_api_key, f"MyGlossary_{info['code']}", "DE", info["code"], entries)
    glossary_ids[info["code"]] = glossary_id
    print(f"Glossary ID for {language}: {glossary_id}")

# Function to translate text using a specific glossary ID
def translate_text_with_glossary(text, target_lang, api_key, glossary_id):
    url = "https://api-free.deepl.com/v2/translate"
    data = {
        'text': text,
        'target_lang': target_lang,
        'source_lang': "DE",  # Specify the source language as German
        'glossary_id': glossary_id
    }
    headers = {'Authorization': f'DeepL-Auth-Key {api_key}'}
    response = requests.post(url, data=data, headers=headers)
    if response.status_code != 200:
        print(f"Error: Received status code {response.status_code}")
        print("Response content:", response.json())
        return "Translation Error"
    return response.json()["translations"][0]["text"]

def translate_excel(file):
    workbook = load_workbook(filename=BytesIO(file.read()))
    def update_cell(sheet, cell_ref, new_value):
        cell = sheet[cell_ref]
        old_fill = cell.fill.copy()
        cell.value = new_value
        cell.alignment = Alignment(wrap_text=True)
        cell.fill = old_fill

    sheet = workbook['DeepL']
    for row in range(2, sheet.max_row + 1):
        german_text_cell = sheet['A{}'.format(row)]
        german_text = german_text_cell.value
        if german_text:
            for language, info in deepl_languages.items():
                target_cell_ref = '{}{}'.format(info["column"], row)
                if not sheet[target_cell_ref].value:
                    glossary_id = glossary_ids.get(info["code"])
                    translated_text = translate_text_with_glossary(
                        german_text, info["code"], deepl_api_key, glossary_id
                    )
                    update_cell(sheet, target_cell_ref, translated_text)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def main():
    st.title("Excel File Translator Tool with Glossary Support (DeepL based)")

    st.markdown("""
        This tool translates German text in an Excel file to English, French, Dutch, Swedish, Czech, and Slovak using DeepL, with glossary support for consistent terminology.
        \nThe sheet must start with the German column and follow this column order: German, English, French, Dutch, Swedish, Czech, Slovak. 
        The German column should be filled with the texts to translate. The exact names of the columns are not important, just the order.
    """)

    uploaded_file = st.file_uploader("Upload Excel file", type=['xlsx'])

    if uploaded_file is not None:
        if st.button('Translate Excel File'):
            with st.spinner('Translating...'):
                output_file = translate_excel(uploaded_file)
                st.success('Translation Completed!')
                st.download_button('Download Translated File', 
                                   data=output_file, 
                                   file_name='translated_excel.xlsx',
                                   mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    main()
