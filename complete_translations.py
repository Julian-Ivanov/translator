from openpyxl import load_workbook
import requests
import uuid
from openpyxl.styles import Alignment

# DeepL API key
deepl_api_key = '25f727f2-4270-6944-171f-da78e41dcef0:fx'

# Microsoft Translator API key
microsoft_translator_api_key = '909ee48eb983487bb43a2dc397ff64e9'

# DeepL languages and their language codes
deepl_languages = {
    "English": {"code": "EN", "column": "B"},
    "French": {"code": "FR", "column": "C"},
    "Dutch": {"code": "NL", "column": "D"},
    "Swedish": {"code": "SV", "column": "E"},
    "Czech": {"code": "CS", "column": "F"},
    "Slovak": {"code": "SK", "column": "G"},
}

# Microsoft Translator languages and their language codes
microsoft_languages = {
    "English": {"code": "en", "column": "B"},
    "French": {"code": "fr", "column": "C"},
    "Dutch": {"code": "nl", "column": "D"},
    "Swedish": {"code": "sv", "column": "E"},
    "Czech": {"code": "cs", "column": "F"},
    "Slovak": {"code": "sk", "column": "G"},
}

# Function to translate text using DeepL
def translate_text_deepl(text_to_translate, target_language_code, api_key):
    url = "https://api-free.deepl.com/v2/translate"
    data = {
        'text': text_to_translate,
        'target_lang': target_language_code
    }
    headers = {
        'Authorization': f'DeepL-Auth-Key {api_key}'
    }
    response = requests.post(url, data=data, headers=headers)
    
    if response.status_code != 200:
        print(f"Error: Received status code {response.status_code}")
        return "Translation Error"
    
    translation_result = response.json()
    if "translations" in translation_result:
        return translation_result["translations"][0]["text"]
    else:
        return "Translation Error"

# Function to translate text using Microsoft Translator
def translate_text_microsoft(text_to_translate, target_language_code, api_key):
    endpoint = "https://api.cognitive.microsofttranslator.com"
    path = "/translate?api-version=3.0"
    params = "&to=" + target_language_code
    constructed_url = endpoint + path + params

    headers = {
        'Ocp-Apim-Subscription-Key': api_key,
        'Ocp-Apim-Subscription-Region': 'germanywestcentral', 
        'Content-type': 'application/json',
        'X-ClientTraceId': str(uuid.uuid4())
    }

    body = [{'text': text_to_translate}]
    response = requests.post(constructed_url, headers=headers, json=body)

    if response.status_code != 200:
        print(f"Error: Received status code {response.status_code}")
        return "Translation Error!"

    translation_result = response.json()
    if translation_result:
        return translation_result[0]["translations"][0]["text"]
    else:
        return "Translation Error!"

# Load the workbook
workbook = load_workbook(filename='translations.xlsx')

# Function to update a cell with translation while preserving formatting
def update_cell(sheet, cell_ref, new_value):
    cell = sheet[cell_ref]

    # Store old formatting properties
    old_fill = cell.fill.copy()

    # Update the cell value
    cell.value = new_value

    # Reapply the old fill properties and enable wrap text
    cell.fill = old_fill
    cell.alignment = Alignment(wrap_text=True)

# Process DeepL and Microsoft Translator sheets
for sheet_name in ['DeepL', 'Microsoft Translator']:
    sheet = workbook[sheet_name]
    for row in range(2, sheet.max_row + 1):
        german_text_cell = sheet['A{}'.format(row)]
        german_text = german_text_cell.value
        if german_text:
            languages = deepl_languages if sheet_name == 'DeepL' else microsoft_languages
            for language, info in languages.items():
                target_cell_ref = '{}{}'.format(info["column"], row)
                if not sheet[target_cell_ref].value:
                    translate_func = translate_text_deepl if sheet_name == 'DeepL' else translate_text_microsoft
                    translated_text = translate_func(german_text, info["code"], deepl_api_key if sheet_name == 'DeepL' else microsoft_translator_api_key)
                    update_cell(sheet, target_cell_ref, translated_text)

# Save the workbook with the original format preserved
workbook.save(filename='updated_translations.xlsx')

print("Die Texte wurden mit DeepL und Microsoft Translator übersetzt")
